"""
Generate BBL GIS IMMO Excel from DATAMODEL.json (the canonical source of truth).
Matches the color scheme of the original BBL GIS IMMO 01-02-2025.xlsx.

Usage:
    python docs/generate_excel.py

Reads:  docs/DATAMODEL.json
Writes: docs/BBL GIS IMMO 18-03-2026.xlsx
"""

import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Color scheme (matches original theme colors converted to RGB)
# ---------------------------------------------------------------------------
FILLS = {
    "internal_light": PatternFill("solid", fgColor="DEEBF6"),   # theme8 t0.8
    "internal_med":   PatternFill("solid", fgColor="BDD7EE"),   # theme8 t0.6
    "financial":      PatternFill("solid", fgColor="F7CBAC"),   # theme5 t0.6
    "people":         PatternFill("solid", fgColor="9CC3E5"),   # theme8 t0.4
    "address":        PatternFill("solid", fgColor="FBE5D5"),   # theme5 t0.8
    "coordinates":    PatternFill("solid", fgColor="E7E6E6"),   # theme2 t0.0
    "survey":         PatternFill("solid", fgColor="D8D8D8"),   # theme0 t-0.15
    "zoning":         PatternFill("solid", fgColor="FFF2CC"),   # theme7 t0.8
    "heritage":       PatternFill("solid", fgColor="E2EEDA"),   # theme9 t0.8
    "dimensions":     PatternFill("solid", fgColor="FBE5D5"),   # theme5 t0.8
    "other":          PatternFill("solid", fgColor="F2F2F2"),   # theme0 t-0.05
    "dev":            PatternFill("solid", fgColor="ED7D31"),   # theme5 t0.0
    "project":        PatternFill("solid", fgColor="BDD7EE"),   # same as internal_med
}


def get_fill(entity_id, group, status, field=None):
    """Determine row fill color based on entity, group, status, and field."""
    if status == "DEV":
        return FILLS["dev"]

    if entity_id == "building":
        if group == "Internal Master Data":
            if field in ("bbl_awrt", "bbl_bwrt"):
                return FILLS["financial"]
            if field in ("bbl_ovtw", "bbl_pvtw"):
                return FILLS["people"]
            return FILLS["internal_light"]
        if group == "Internal Master Data 2":
            if field in ("bbl_awrt", "bbl_bwrt"):
                return FILLS["financial"]
            if field in ("bbl_ovtw", "bbl_pvtw"):
                return FILLS["people"]
            return FILLS["internal_med"]

    if entity_id == "parcel":
        if group == "Internal Master Data":
            if field in ("bbl_awrt", "bbl_bwrt"):
                return FILLS["financial"]
            return FILLS["internal_light"]

    # Shared group colors
    group_map = {
        "Internal Master Data": "internal_light",
        "Location": "address",
        "Official Survey": "survey",
        "Zoning": "zoning",
        "Heritage Protection": "heritage",
        "Dimensions - SIA 416": "dimensions",
        "Dimensions - SIA 380": "dimensions",
        "Dimensions - eBKP-H": "dimensions",
        "Other Dimensions": "dimensions",
        "Other": "other",
        "Project Controlling": "project",
        "Cleaning": "dev",
        "Workplaces": "dev",
    }
    return FILLS.get(group_map.get(group, "other"), FILLS["other"])


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------
HEADERS = [
    "ID", "Status", "Anwendung", "Geschaeftsobjekt", "Merkmal Gruppe",
    "Merkmal Alias DE", "Merkmal Alias EN", "Merkmal DB", "Merkmal DB Len",
    "Format", "Werteliste", "Herkunft", "Beschreibung DE", "Key",
    "Beschreibung EN", "Sichtbar",
]

COL_WIDTHS = {
    "A": 6, "B": 8, "C": 16, "D": 22, "E": 30, "F": 35, "G": 40,
    "H": 16, "I": 14, "J": 10, "K": 16, "L": 28, "M": 60, "N": 6,
    "O": 60, "P": 10,
}

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
DATA_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def load_datamodel():
    """Load DATAMODEL.json from the docs directory."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, "DATAMODEL.json")
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def main():
    model = load_datamodel()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attribute"

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = THIN_BORDER

    # Set column widths
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Write data rows from DATAMODEL.json
    row_idx = 2
    total_attrs = 0

    for entity in model["entities"]:
        entity_id = entity["id"]
        entity_name = entity["name_en"]

        for attr in entity["attributes"]:
            field = attr["field"]
            field_len = len(field)
            visible = "Ja" if attr.get("visible", False) else ""

            values = [
                attr["sort"],
                attr["status"],
                model["system"],
                entity_name,
                attr["group"],
                attr["alias_de"],
                attr["alias_en"],
                field,
                field_len,
                attr["format"],
                attr.get("value_list", ""),
                attr.get("source", ""),
                attr["description_de"],
                attr.get("key", ""),
                attr["description_en"],
                visible,
            ]

            fill = get_fill(entity_id, attr["group"], attr["status"], field)

            for col_idx_inner, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx_inner, value=value)
                cell.font = DATA_FONT
                cell.fill = fill
                cell.border = THIN_BORDER
                if col_idx_inner in (1, 2, 9, 14, 16):  # ID, Status, Len, Key, Visible
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            row_idx += 1
            total_attrs += 1

    # Add Excel table
    last_row = row_idx - 1
    last_col = get_column_letter(len(HEADERS))
    table_ref = f"A1:{last_col}{last_row}"
    table = Table(displayName="AttributeTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Add autofilter
    ws.auto_filter.ref = table_ref

    # Create empty Werteliste sheet
    ws2 = wb.create_sheet("Werteliste")
    ws2.cell(row=1, column=1, value="Value List")
    ws2.cell(row=1, column=2, value="Code")
    ws2.cell(row=1, column=3, value="Description EN")
    ws2.cell(row=1, column=4, value="Description DE")
    for vl_col in range(1, 5):
        cell = ws2.cell(row=1, column=vl_col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Save
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "BBL GIS IMMO 18-03-2026.xlsx")
    wb.save(output_path)
    print(f"Saved to {output_path}")
    print(f"  {total_attrs} attribute rows written")
    print(f"  {len(HEADERS)} columns")
    print(f"  {len(model['entities'])} entities")


if __name__ == "__main__":
    main()
